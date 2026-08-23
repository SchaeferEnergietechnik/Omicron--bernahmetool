"""Windows worker for unattended Omicron OCC exports and Excel refreshes.

The worker deliberately has no interactive prompts. The GUI must resolve all
conflicts before it writes the job file. Progress is emitted as JSON Lines to
stdout so a separate GUI process can display it and request cancellation.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pywinauto import Application
from pywinauto.findwindows import ElementNotFoundError, find_windows
import win32com.client


WINDOW_TITLE_RE = ".*OMICRON Control Center.*"
EXPORT_DIALOG_RE = ".*Datenexport.*"
EXCEL_SHEET = "Prüfprotokoll"
MACRO_PROTOCOL_NO = "Tabelle1.Protokollnummer_generieren_unsichtbar"
MACRO_TOGGLE_SECTIONS_CANDIDATES = [
    "BereicheEinOderAusblenden_Start",
    "Modul1.BereicheEinOderAusblenden_Start",
]
MACRO_HIDE_EMPTY_ROWS = "Tabelle7.ZeilenAusblendenWennLeer"
DEFAULT_EXPORT_DIRECTORY = Path(r"C:\Omicron_Datenexport")
TERMINEXCEL_PATH = Path(r"Y:\GES Energietechnik\Termine.xlsx")
TERMINEXCEL_SHEET = "Termine"
CHECKLIST_SHEET = "Schutzprüf-Checkliste"
GENERAL_SHEET = "Allgemeine Angaben"
CUSTOMER_LIST_SHEET = "Kunden"
CUSTOMER_TARGET_CELL = "C2"
INSPECTOR_DATE_CELL = "B7"

INSPECTOR_ALIASES: dict[str, str] = {
    "helmchen": "Niklas Helmchen",
    "niklas helmchen": "Niklas Helmchen",
    "n helmchen": "Niklas Helmchen",
    "n. helmchen": "Niklas Helmchen",
    "faethke": "Pascal Fäthke",
    "fathke": "Pascal Fäthke",
    "fäthke": "Pascal Fäthke",
    "pascal fäthke": "Pascal Fäthke",
    "p. fäthke": "Pascal Fäthke",
    "schmidt": "Hagen Schmidt",
    "hagen schmidt": "Hagen Schmidt",
    "h. schmidt": "Hagen Schmidt",
    "koehn": "Kevin Koehn",
    "kevin koehn": "Kevin Koehn",
    "k. koehn": "Kevin Koehn",
    "wendt": "Sebastian Wendt",
    "sebastian wendt": "Sebastian Wendt",
    "s. wendt": "Sebastian Wendt",
    "mummhardt": "Elias Mummhardt",
    "elias mummhardt": "Elias Mummhardt",
    "e. mummhardt": "Elias Mummhardt",
    "kolzer": "Finn Kolzer",
    "finn kolzer": "Finn Kolzer",
    "f. kolzer": "Finn Kolzer",
}

INSPECTOR_ALIAS_ITEMS_SORTED = sorted(INSPECTOR_ALIASES.items(), key=lambda item: len(item[0]), reverse=True)

IGNORED_INSPECTOR_ALIASES = (
    "schaefer",
    "schäfer",
    "g schafer",
    "g. schafer",
    "g schäfer",
    "g. schäfer",
    "mundkowski",
    "t mundkowski",
    "t. mundkowski",
)

INTERNAL_TERM_KEYWORDS = (
    "urlaub",
    "elternzeit",
    "intern",
    "ges intern",
    "schulung",
    "krank",
    "büro",
    "buero",
    "homeoffice",
)

CUSTOMER_MATCH_STOPWORDS = {
    "gmbh",
    "mbh",
    "ag",
    "kg",
    "kgaa",
    "gbr",
    "ug",
    "haftungsbeschraenkt",
    "co",
    "und",
    "the",
    "der",
    "die",
    "das",
    "ein",
    "eine",
    "einer",
    "einem",
    "einen",
    "&",
}


class CancellationRequested(Exception):
    """Raised at safe boundaries after the GUI writes the cancellation file."""


class Worker:
    def __init__(self, job: dict[str, Any], cancel_file: Path | None):
        self.job = job
        self.cancel_file = cancel_file
        self.run_started_monotonic: float | None = None
        self.failures: list[dict[str, str]] = []
        self.skipped: list[dict[str, str]] = []
        self.archive_warnings: list[dict[str, str]] = []
        self.skip_section_macro = bool(job.get("skipSectionMacro", False))

    def emit(self, event: str, **payload: Any) -> None:
        print(json.dumps({"timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"), "event": event, **payload}, ensure_ascii=False), flush=True)

    def check_cancelled(self) -> None:
        if self.cancel_file and self.cancel_file.exists():
            raise CancellationRequested()

    def wait(self, seconds: float) -> None:
        end_time = time.monotonic() + seconds
        while time.monotonic() < end_time:
            self.check_cancelled()
            time.sleep(min(0.5, max(0, end_time - time.monotonic())))

    def wait_for_main_window(self, timeout: int = 30):
        end_time = time.monotonic() + timeout
        last_error: Exception | None = None
        while time.monotonic() < end_time:
            self.check_cancelled()
            try:
                windows = find_windows(title_re=WINDOW_TITLE_RE)
                if windows:
                    app = Application(backend="uia").connect(handle=windows[0])
                    window = app.window(handle=windows[0])
                    window.wait("visible", timeout=5)
                    window.set_focus()
                    return window
            except Exception as error:  # pywinauto errors vary by Omicron version
                last_error = error
            self.wait(1)
        raise ElementNotFoundError(f"Kein OMICRON-Fenster gefunden: {last_error}")

    def open_export_dialog(self, window) -> bool:
        try:
            window.menu_select("Datei->Daten exportieren...")
            self.wait(1)
            return True
        except Exception:
            pass
        
        # Fallback: Try clicking Datei menu with different strategies
        for candidate in (
            {"title": "Datei", "control_type": "TabItem"},
            {"title": "Datei", "control_type": "Button"},
            {"title": "Datei"},
        ):
            try:
                control = window.child_window(**candidate)
                if control.exists(timeout=2):
                    control.click_input()
                    self.wait(1)
                    break
            except Exception:
                continue
        else:
            # Fallback: Try keyboard shortcut Alt+D
            try:
                window.set_focus()
                window.type_keys("%D")
                self.wait(1)
                return self._click_export_menu_item(window)
            except Exception:
                return False
        
        return self._click_export_menu_item(window)
    
    def _click_export_menu_item(self, window) -> bool:
        for candidate in (
            {"title": "Daten exportieren...", "control_type": "MenuItem"},
            {"title": "Daten exportieren...", "control_type": "ListItem"},
            {"title": "Daten exportieren..."},
            {"title": "Daten exportieren"},
        ):
            try:
                control = window.child_window(**candidate)
                if control.exists(timeout=2):
                    control.click_input()
                    self.wait(1)
                    return True
            except Exception:
                continue
        
        # Fallback: Try keyboard navigation
        try:
            window.type_keys("{DOWN 8}{ENTER}")
            self.wait(1)
            return True
        except Exception:
            return False

    def confirm_export_dialog(self, timeout: int = 30) -> bool:
        end_time = time.monotonic() + timeout
        while time.monotonic() < end_time:
            self.check_cancelled()
            try:
                windows = find_windows(title_re=EXPORT_DIALOG_RE)
                if windows:
                    app = Application(backend="uia").connect(handle=windows[0])
                    dialog = app.window(handle=windows[0])
                    dialog.wait("visible", timeout=5)
                    dialog.set_focus()
                    for title in ("OK", "Ok", "ok"):
                        button = dialog.child_window(title=title, control_type="Button")
                        if button.exists(timeout=1):
                            button.click_input()
                            self.wait(1)
                            return True
                    dialog.type_keys("{ENTER}")
                    self.wait(1)
                    return True
            except Exception:
                pass
            self.wait(1)
        return False

    def close_omicron(self, window) -> None:
        try:
            window.close()
            self.wait(2)
        except CancellationRequested:
            raise
        except Exception:
            return

        # Closing an OCC without changes can still show a save question. Never wait
        # for a user: discard instead, otherwise record a later file failure.
        try:
            for handle in find_windows(title_re=WINDOW_TITLE_RE):
                dialog = Application(backend="uia").connect(handle=handle).window(handle=handle)
                text = " ".join(item.window_text() for item in dialog.descendants() if item.window_text()).lower()
                if "speichern" in text:
                    for title in ("Nein", "NEIN", "No"):
                        button = dialog.child_window(title=title, control_type="Button")
                        if button.exists(timeout=1):
                            button.click_input()
                            return
        except Exception:
            return

    def terminate_mashup_loader(self) -> None:
        """Stoppt den Power-Query-Mashup-Loader vor einem neuen Verarbeitungslauf."""
        try:
            result = subprocess.run(
                ["taskkill", "/IM", "Microsoft.Mashup.Container.Loader.exe", "/F"],
                capture_output=True,
                text=True,
                shell=False,
            )
            if result.returncode == 0:
                self.emit("mashup_terminated")
            else:
                # Kein Fehlerfall: Prozess lief ggf. nicht.
                self.emit("mashup_not_running")
        except Exception as error:
            self.emit("mashup_termination_failed", message=str(error))

    def export_directory(self) -> Path:
        return Path(self.job.get("exportDirectory", DEFAULT_EXPORT_DIRECTORY))

    def csv_snapshot(self, directory: Path) -> dict[Path, tuple[int, int]]:
        try:
            return {
                file_path: (file_path.stat().st_mtime_ns, file_path.stat().st_size)
                for file_path in directory.glob("*.csv")
                if file_path.is_file()
            }
        except OSError:
            return {}

    def wait_for_export_output(
        self,
        directory: Path,
        before_export: dict[Path, tuple[int, int]],
        started_at_ns: int,
        timeout: int = 120,
    ) -> list[Path]:
        """Bestätigt den Export erst, wenn neue oder geänderte CSV-Dateien stabil sind."""
        self.emit("occ_export_waiting", exportDirectory=str(directory), timeoutSeconds=timeout)
        deadline = time.monotonic() + timeout
        stable_snapshot: dict[Path, tuple[int, int]] | None = None
        stable_since: float | None = None

        while time.monotonic() < deadline:
            self.check_cancelled()
            current = self.csv_snapshot(directory)
            changed_files = [
                path for path, signature in current.items()
                if before_export.get(path) != signature or signature[0] >= started_at_ns
            ]
            if changed_files:
                changed_snapshot = {path: current[path] for path in changed_files}
                if changed_snapshot == stable_snapshot:
                    if stable_since is not None and time.monotonic() - stable_since >= 2:
                        self.emit("occ_export_data_ready", csvFiles=[str(path) for path in changed_files])
                        return changed_files
                else:
                    stable_snapshot = changed_snapshot
                    stable_since = time.monotonic()
            else:
                stable_snapshot = None
                stable_since = None
            self.wait(1)

        raise TimeoutError(
            f"Keine neuen oder aktualisierten CSV-Dateien in {directory} nach {timeout} Sekunden erkannt"
        )

    def export_occ(self, occ_path: Path) -> None:
        if not occ_path.is_file():
            raise FileNotFoundError(occ_path)
        self.check_cancelled()
        window = None
        last_error: Exception | None = None
        for attempt in range(1, 4):
            self.check_cancelled()
            try:
                os.startfile(str(occ_path))
            except Exception as error:
                last_error = error
                if attempt == 3:
                    raise
                self.wait(2)
                continue

            self.wait(5 if attempt == 1 else 8)
            try:
                window = self.wait_for_main_window(timeout=45)
                break
            except ElementNotFoundError as error:
                last_error = error
                if attempt == 3:
                    raise
                self.emit("occ_open_retry", occPath=str(occ_path), attempt=attempt)
                self.terminate_mashup_loader()
                self.wait(2)

        if window is None:
            raise RuntimeError(f"OCC-Datei konnte nicht geoeffnet werden: {last_error}")
        try:
            if not self.open_export_dialog(window):
                raise RuntimeError("Exportmenü konnte nicht geöffnet werden")
            self.wait(2)
            export_directory = self.export_directory()
            before_export = self.csv_snapshot(export_directory)
            export_started_at_ns = time.time_ns()
            if not self.confirm_export_dialog():
                raise RuntimeError("Exportdialog konnte nicht bestätigt werden")
            self.wait_for_export_output(export_directory, before_export, export_started_at_ns)
        finally:
            self.close_omicron(window)

    def elapsed_seconds(self) -> int:
        if self.run_started_monotonic is None:
            return 0
        return max(0, int(time.monotonic() - self.run_started_monotonic))

    def write_report_if_needed(self, succeeded: int, failed: int, skipped: int) -> str | None:
        report_path_value = self.job.get("reportPath")
        if not report_path_value:
            return None
        report_path = Path(report_path_value)
        payload = {
            "timestamp": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "elapsedSeconds": self.elapsed_seconds(),
            "summary": {
                "succeeded": succeeded,
                "failed": failed,
                "skipped": skipped,
                "archiveWarnings": len(self.archive_warnings),
            },
            "failures": self.failures,
            "skippedItems": self.skipped,
            "archiveWarnings": self.archive_warnings,
        }
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return str(report_path)

    def archive_root_directory(self) -> Path | None:
        value = self.job.get("archiveRoot")
        if not value:
            return None
        return Path(str(value))

    def relative_folder_path(self, item: dict[str, Any], source_dir: Path) -> Path:
        relative_value = str(item.get("folderRelativePath", "")).strip()
        if relative_value:
            normalized = relative_value.replace("\\", "/").strip("/")
            parts = [part for part in normalized.split("/") if part and part != "."]
            if parts:
                return Path(*parts)
        return Path(source_dir.name)

    def archive_processed_folder(self, item: dict[str, Any]) -> bool:
        archive_root = self.archive_root_directory()
        working_dir_value = item.get("workingDirectory")
        item_id = str(item.get("id", ""))

        if archive_root is None or not working_dir_value:
            warning = "Archivierung übersprungen: archiveRoot oder workingDirectory fehlt"
            self.emit("folder_archive_skipped", itemId=item_id, message=warning)
            self.archive_warnings.append({"itemId": item_id, "message": warning})
            return False

        source_dir = Path(str(working_dir_value))
        if not source_dir.is_dir():
            warning = f"Archivierung übersprungen: Fundordner nicht gefunden ({source_dir})"
            self.emit("folder_archive_skipped", itemId=item_id, sourcePath=str(source_dir), message=warning)
            self.archive_warnings.append({"itemId": item_id, "sourcePath": str(source_dir), "message": warning})
            return False

        target_dir = archive_root / self.relative_folder_path(item, source_dir)
        if target_dir.exists():
            warning = f"Zielordner existiert bereits, keine Überschreibung: {target_dir}"
            self.emit(
                "folder_archive_conflict",
                itemId=item_id,
                sourcePath=str(source_dir),
                targetPath=str(target_dir),
                message=warning,
            )
            self.archive_warnings.append(
                {
                    "itemId": item_id,
                    "sourcePath": str(source_dir),
                    "targetPath": str(target_dir),
                    "message": warning,
                }
            )
            return False

        try:
            target_dir.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(source_dir), str(target_dir))
            self.emit(
                "folder_archived",
                itemId=item_id,
                sourcePath=str(source_dir),
                targetPath=str(target_dir),
            )
            return True
        except Exception as error:
            warning = f"Archivierung fehlgeschlagen: {error}"
            self.emit(
                "folder_archive_failed",
                itemId=item_id,
                sourcePath=str(source_dir),
                targetPath=str(target_dir),
                message=warning,
            )
            self.archive_warnings.append(
                {
                    "itemId": item_id,
                    "sourcePath": str(source_dir),
                    "targetPath": str(target_dir),
                    "message": warning,
                }
            )
            return False

    def sort_occ_paths(self, occ_paths: list[Path]) -> list[Path]:
        """Sortiert OCC-Dateien stabil: zuerst nicht-EZE (z. B. NAP), dann EZE."""
        def sort_key(path: Path) -> tuple[int, str]:
            name = path.name.lower()
            is_eze = "eze" in name
            return (1 if is_eze else 0, name)

        return sorted(occ_paths, key=sort_key)

    def resolve_excel_groups(self, item: dict[str, Any]) -> list[tuple[Path, list[Path]]]:
        """Liefert stabile Gruppen: [(excel_path, [occ_paths...])]."""
        mappings = item.get("mappings")
        if mappings:
            grouped: dict[Path, list[Path]] = {}
            for mapping in mappings:
                occ_value = mapping.get("occPath")
                excel_value = mapping.get("excelPath")
                if not occ_value or not excel_value:
                    raise RuntimeError("Ungültige manuelle Zuordnung: OCC- oder Excel-Pfad fehlt")
                excel_path = Path(excel_value)
                occ_path = Path(occ_value)
                grouped.setdefault(excel_path, []).append(occ_path)

            groups: list[tuple[Path, list[Path]]] = []
            for excel_path in sorted(grouped.keys(), key=lambda path: path.name.lower()):
                groups.append((excel_path, self.sort_occ_paths(grouped[excel_path])))
            return groups

        occ_paths = self.sort_occ_paths([Path(path) for path in item.get("occPaths", [])])
        excel_path = Path(item["excelPath"])
        return [(excel_path, occ_paths)]

    def run_macro(self, excel, workbook_name: str, macro_name: str) -> None:
        excel.Application.Run(f"'{workbook_name}'!{macro_name}")
        self.wait(2)

    def normalize_text(self, value: str) -> str:
        normalized = value.lower().strip()
        normalized = normalized.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
        normalized = re.sub(r"\s+", " ", normalized)
        return normalized

    def excel_value_to_date(self, value: Any) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            try:
                return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
            except Exception:
                return None
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    def map_inspector_name(self, raw_name: str) -> str | None:
        normalized = self.normalize_text(raw_name)
        if normalized in INSPECTOR_ALIASES:
            return INSPECTOR_ALIASES[normalized]
        for alias, full_name in INSPECTOR_ALIAS_ITEMS_SORTED:
            pattern = rf"(?<![a-z0-9]){re.escape(alias)}(?![a-z0-9])"
            if re.search(pattern, normalized):
                return full_name
        return None

    def is_ignored_inspector_label(self, raw_label: str) -> bool:
        normalized = self.normalize_text(raw_label)
        for alias in IGNORED_INSPECTOR_ALIASES:
            pattern = rf"(?<![a-z0-9]){re.escape(alias)}(?![a-z0-9])"
            if re.search(pattern, normalized):
                return True
        return False

    def extract_inspector_names(self, raw_label: str) -> list[str]:
        normalized = self.normalize_text(raw_label)
        if not normalized:
            return []
        if self.is_ignored_inspector_label(raw_label):
            return []

        if normalized in INSPECTOR_ALIASES:
            return [INSPECTOR_ALIASES[normalized]]

        matches: list[str] = []
        for alias, full_name in INSPECTOR_ALIAS_ITEMS_SORTED:
            pattern = rf"(?<![a-z0-9]){re.escape(alias)}(?![a-z0-9])"
            if re.search(pattern, normalized):
                matches.append(full_name)
        return list(dict.fromkeys(matches))

    def map_checked_labels_to_unique_inspectors(self, labels: list[str]) -> list[str]:
        resolved: list[str] = []
        for label in labels:
            candidates = self.extract_inspector_names(label)
            # Nur eindeutig zuordenbare Labels berücksichtigen. Mehrdeutige Sammeltexte
            # (z. B. eine ganze Namenszeile als Caption) würden sonst fälschlich mehrere
            # Prüfer als "angehakt" markieren.
            if len(candidates) == 1:
                resolved.append(candidates[0])
        return list(dict.fromkeys(resolved))

    def checked_inspector_from_checkboxes(self, workbook) -> str:
        sheet = workbook.Worksheets(CHECKLIST_SHEET)
        forms_checked_labels: list[str] = []
        activex_checked_labels: list[str] = []

        # Forms-Checkboxen
        try:
            checkboxes = sheet.CheckBoxes()
            for index in range(1, int(checkboxes.Count) + 1):
                checkbox = checkboxes.Item(index)
                try:
                    value = int(checkbox.Value)
                except Exception:
                    value = 0
                if value == 1:
                    label = ""
                    try:
                        label = str(checkbox.Caption).strip()
                    except Exception:
                        pass
                    if label:
                        forms_checked_labels.append(label)
        except Exception:
            pass

        # ActiveX-Checkboxen
        try:
            ole_objects = sheet.OLEObjects()
            for index in range(1, int(ole_objects.Count) + 1):
                ole = ole_objects.Item(index)
                try:
                    prog_id = str(ole.progID)
                except Exception:
                    prog_id = ""
                if "CheckBox" not in prog_id:
                    continue
                try:
                    value = bool(ole.Object.Value)
                except Exception:
                    value = False
                if value:
                    label = ""
                    try:
                        label = str(ole.Object.Caption).strip()
                    except Exception:
                        pass
                    if label:
                        activex_checked_labels.append(label)
        except Exception:
            pass

        forms_names = self.map_checked_labels_to_unique_inspectors(forms_checked_labels)
        activex_names = self.map_checked_labels_to_unique_inspectors(activex_checked_labels)
        combined_names = list(dict.fromkeys(forms_names + activex_names))
        raw_checked = list(dict.fromkeys(forms_checked_labels + activex_checked_labels))
        ignored_checked = [label for label in raw_checked if self.is_ignored_inspector_label(label)]

        # Einige Vorlagen enthalten sowohl Forms- als auch ActiveX-Elemente.
        # Wenn nur eine Quelle konsistent genau einen Prüfer liefert, akzeptieren wir sie.
        if len(activex_names) == 1 and (not forms_names or forms_names == activex_names):
            return activex_names[0]
        if len(forms_names) == 1 and (not activex_names or activex_names == forms_names):
            return forms_names[0]
        if len(combined_names) == 1:
            return combined_names[0]

        if len(combined_names) > 1:
            raise RuntimeError(
                "Mehrere angehakte Prüfer erkannt: "
                f"Forms={forms_names or forms_checked_labels}, ActiveX={activex_names or activex_checked_labels}"
            )

        if ignored_checked and len(ignored_checked) == len(raw_checked):
            raise RuntimeError(
                "Nur veraltete Prüferbezeichnungen angehakt (nicht unterstützt): "
                + ", ".join(ignored_checked)
            )

        if raw_checked:
            raise RuntimeError(
                "Angehakte Prüfer-Checkbox gefunden, aber kein unterstützter Prüfername erkannt: "
                + ", ".join(raw_checked)
            )
        raise RuntimeError("Kein angehakter Prüfer in Schutzprüf-Checkliste gefunden")

    def read_customer_list(self, workbook) -> list[str]:
        customer_sheet = workbook.Worksheets(CUSTOMER_LIST_SHEET)
        values = customer_sheet.Range("A1:A35").Value
        customers: list[str] = []
        for row in values:
            value = row[0] if isinstance(row, tuple) else row
            if value is None:
                continue
            text = str(value).strip()
            if text:
                customers.append(text)
        return customers

    def contains_internal_keyword(self, value: str) -> bool:
        normalized = self.normalize_text(value)
        return any(keyword in normalized for keyword in INTERNAL_TERM_KEYWORDS)

    def customer_keywords(self, value: str) -> set[str]:
        normalized = self.normalize_text(value)
        tokens = re.findall(r"[a-z0-9]+", normalized)
        return {
            token
            for token in tokens
            if len(token) >= 3 and token not in CUSTOMER_MATCH_STOPWORDS
        }

    def matching_customers_from_list(self, raw_customer: str, customers: list[str]) -> list[str]:
        normalized_raw = self.normalize_text(raw_customer)
        raw_keywords = self.customer_keywords(raw_customer)
        matches: list[str] = []

        for customer in customers:
            normalized_customer = self.normalize_text(customer)
            if normalized_customer == normalized_raw:
                matches.append(customer)
                continue

            customer_keywords = self.customer_keywords(customer)
            if raw_keywords and customer_keywords and raw_keywords.intersection(customer_keywords):
                matches.append(customer)

        return matches

    def find_inspector_header_column(self, header_row: tuple[Any, ...], inspector_full_name: str) -> int:
        normalized_target = self.normalize_text(inspector_full_name)
        for index, value in enumerate(header_row, start=1):
            if value is None:
                continue
            if self.normalize_text(str(value)) == normalized_target:
                return index
        raise RuntimeError(f"Prüferblock '{inspector_full_name}' im Blatt '{TERMINEXCEL_SHEET}' nicht gefunden")

    def resolve_customer_from_terminexcel(self, excel, workbook, item: dict[str, Any]) -> str:
        if not TERMINEXCEL_PATH.is_file():
            raise RuntimeError(f"Terminexcel nicht erreichbar: {TERMINEXCEL_PATH}")

        inspector_full_name = self.checked_inspector_from_checkboxes(workbook)
        self.emit("inspector_detected", itemId=item.get("id"), inspector=inspector_full_name)

        checklist_sheet = workbook.Worksheets(CHECKLIST_SHEET)
        exam_date = self.excel_value_to_date(checklist_sheet.Range(INSPECTOR_DATE_CELL).Value)
        if exam_date is None:
            raise RuntimeError("Prüfdatum in Schutzprüf-Checkliste!B7 ist leer oder ungültig")

        customers = self.read_customer_list(workbook)
        if not customers:
            raise RuntimeError("Kundenliste Kunden!A1:A35 ist leer")

        # Die Terminexcel wird absichtlich nicht per Excel-UI geöffnet,
        # damit keine zusätzliche sichtbare Arbeitsmappe oder COM-Dialoge entstehen.
        try:
            termin_workbook = load_workbook(filename=str(TERMINEXCEL_PATH), read_only=True, data_only=True)
        except Exception as error:
            raise RuntimeError(f"Terminexcel konnte nicht direkt gelesen werden: {error}") from error

        try:
            if TERMINEXCEL_SHEET not in termin_workbook.sheetnames:
                raise RuntimeError(f"Blatt '{TERMINEXCEL_SHEET}' in Terminexcel nicht gefunden")
            termin_sheet = termin_workbook[TERMINEXCEL_SHEET]
            header_cells = next(termin_sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_cells:
                raise RuntimeError("Terminexcel enthält keine Kopfzeile")

            inspector_col = self.find_inspector_header_column(header_cells, inspector_full_name)
            date_col = inspector_col - 2
            customer_col = inspector_col + 1
            if date_col < 1:
                raise RuntimeError(f"Ungültige Spaltenstruktur für Prüferblock '{inspector_full_name}'")

            raw_customer_candidates: list[str] = []
            for row in termin_sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                row_date = self.excel_value_to_date(row[date_col - 1] if len(row) >= date_col else None)
                if row_date != exam_date:
                    continue
                customer_value = row[customer_col - 1] if len(row) >= customer_col else None
                if customer_value is None:
                    continue
                customer_text = str(customer_value).strip()
                if not customer_text:
                    continue
                if self.contains_internal_keyword(customer_text):
                    continue
                raw_customer_candidates.append(customer_text)

            matched_customer_candidates: list[str] = []
            for raw_customer in raw_customer_candidates:
                matched_customer_candidates.extend(
                    self.matching_customers_from_list(raw_customer, customers)
                )

            # Optionaler manueller Override pro Item (z. B. aus späterer GUI-Auswahl).
            manual_customer = str(item.get("manualCustomer", "")).strip()
            if manual_customer:
                if manual_customer not in customers:
                    raise RuntimeError(f"Manueller Kunde nicht in Kundenliste A1:A35: {manual_customer}")
                return manual_customer

            unique_candidates = list(dict.fromkeys(matched_customer_candidates))
            if len(unique_candidates) == 1:
                return unique_candidates[0]
            if len(unique_candidates) > 1:
                self.emit(
                    "customer_selection_required",
                    itemId=item.get("id"),
                    inspector=inspector_full_name,
                    examDate=str(exam_date),
                    options=unique_candidates,
                )
                raise RuntimeError(
                    "Mehrere passende Kunden gefunden. Manuelle Auswahl erforderlich: "
                    + ", ".join(unique_candidates)
                )

            self.emit(
                "customer_manual_required",
                itemId=item.get("id"),
                inspector=inspector_full_name,
                examDate=str(exam_date),
                candidates=raw_customer_candidates,
            )
            raise RuntimeError(
                "Kein passender Kunde per Teilwortabgleich in Kunden!A1:A35 gefunden. "
                f"Prüfer={inspector_full_name}, Prüfdatum={exam_date}, "
                f"Kandidaten aus Terminexcel={raw_customer_candidates}. "
                "Manuelle Eingabe in Allgemeine Angaben!C2 erforderlich."
            )
        finally:
            termin_workbook.close()

    def sanitize_filename_part(self, value: str) -> str:
        sanitized = re.sub(r'[\\/:*?"<>|\[\]]+', "_", value)
        sanitized = re.sub(r"\s+", " ", sanitized).strip().strip(".")
        return sanitized

    def resolve_project_title(self, workbook, fallback_path: Path) -> str:
        for sheet_name, cell_name in (("Allgemeine Angaben", "C2"), ("Allgemeine Angaben", "B2"), ("Allgemeine Angaben", "D2")):
            try:
                value = workbook.Worksheets(sheet_name).Range(cell_name).Value
                if value is not None and str(value).strip():
                    return str(value).strip()
            except Exception:
                continue
        return fallback_path.stem

    def fallback_title_from_inspector_and_date(self, inspector_name: str | None, exam_date: date | None, source_path: Path) -> str:
        inspector_part = self.sanitize_filename_part(inspector_name or "Unbekannter_Pruefer")
        date_part = exam_date.isoformat() if exam_date else "Unbekanntes_Datum"
        return f"{inspector_part}_{date_part}" if inspector_part else f"{source_path.stem}_{date_part}"

    def build_output_excel_path(self, source_path: Path, project_title: str) -> Path:
        date_part = time.strftime("%Y-%m-%d")
        suffix = source_path.suffix
        # Excel SaveAs scheitert häufig bei sehr langen Pfaden; wir halten den Dateinamen bewusst kurz.
        title = (self.sanitize_filename_part(project_title) or source_path.stem)[:64]
        base_name = f"{title}_{date_part}"
        max_total_path_len = 210
        available_name_len = max_total_path_len - len(str(source_path.parent)) - 1 - len(suffix)
        if available_name_len < 12:
            available_name_len = 12
        if len(base_name) > available_name_len:
            base_name = base_name[:available_name_len].rstrip(" ._")

        candidate = source_path.with_name(f"{base_name}{suffix}")
        if candidate == source_path or not candidate.exists():
            return candidate

        alt_base = f"{base_name[:max(4, available_name_len - 7)]}_{time.strftime('%H%M%S')}"
        alt_base = alt_base[:available_name_len].rstrip(" ._")
        return source_path.with_name(f"{alt_base}{suffix}")

    def refresh_excel(self, excel_path: Path, item: dict[str, Any]) -> Path:
        if not excel_path.is_file():
            raise FileNotFoundError(excel_path)
        excel = None
        workbook = None
        output_path = excel_path
        inspector_for_fallback: str | None = None
        exam_date_for_fallback: date | None = None
        customer_assignment_failed_without_template = False
        template_customer = ""
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = True
            excel.DisplayAlerts = False
            # Keine interaktiven Link-/Aktualisierungsdialoge in Nachtlaeufen anzeigen.
            try:
                excel.AskToUpdateLinks = False
            except Exception:
                pass
            workbook = excel.Workbooks.Open(str(excel_path.resolve()), UpdateLinks=0)
            try:
                workbook.RefreshAll()
                timeout = time.monotonic() + 180
                while time.monotonic() < timeout:
                    self.check_cancelled()
                    if excel.CalculationState == 0:
                        break
                    self.wait(1)
                try:
                    excel.CalculateUntilAsyncQueriesDone()
                except Exception:
                    pass
                self.wait(5)
            except Exception as error:
                self.emit(
                    "excel_refresh_warning",
                    itemId=item.get("id"),
                    message=f"RefreshAll konnte nicht vollständig abgeschlossen werden: {error}",
                )

            try:
                inspector_for_fallback = self.checked_inspector_from_checkboxes(workbook)
            except Exception:
                inspector_for_fallback = None
            try:
                checklist_sheet = workbook.Worksheets(CHECKLIST_SHEET)
                exam_date_for_fallback = self.excel_value_to_date(checklist_sheet.Range(INSPECTOR_DATE_CELL).Value)
            except Exception:
                exam_date_for_fallback = None

            try:
                template_value = workbook.Worksheets(GENERAL_SHEET).Range(CUSTOMER_TARGET_CELL).Value
                template_customer = str(template_value).strip() if template_value is not None else ""
            except Exception:
                template_customer = ""

            try:
                selected_customer = self.resolve_customer_from_terminexcel(excel, workbook, item)
                workbook.Worksheets(GENERAL_SHEET).Range(CUSTOMER_TARGET_CELL).Value = selected_customer
                self.emit("customer_assigned", itemId=item.get("id"), customer=selected_customer)
            except Exception as error:
                self.emit(
                    "customer_assignment_skipped",
                    itemId=item.get("id"),
                    message=str(error),
                    inspector=inspector_for_fallback,
                    examDate=str(exam_date_for_fallback) if exam_date_for_fallback else None,
                )
                if template_customer:
                    self.emit(
                        "customer_assignment_fallback_template",
                        itemId=item.get("id"),
                        customer=template_customer,
                    )
                else:
                    customer_assignment_failed_without_template = True

            try:
                workbook.Worksheets(EXCEL_SHEET).Activate()
            except Exception as error:
                self.emit(
                    "excel_macro_warning",
                    itemId=item.get("id"),
                    sheet=EXCEL_SHEET,
                    message=f"Arbeitsblatt konnte nicht aktiviert werden: {error}",
                )

            try:
                self.run_macro(excel, workbook.Name, MACRO_PROTOCOL_NO)
            except Exception as error:
                self.emit(
                    "excel_macro_warning",
                    itemId=item.get("id"),
                    macro=MACRO_PROTOCOL_NO,
                    message=f"Makro fehlgeschlagen: {error}",
                )

            if self.skip_section_macro:
                self.emit(
                    "excel_macro_skipped",
                    macroCandidates=MACRO_TOGGLE_SECTIONS_CANDIDATES,
                    reason="Per Benutzeroption übersprungen",
                )
            else:
                last_error: Exception | None = None
                for macro_name in MACRO_TOGGLE_SECTIONS_CANDIDATES:
                    try:
                        self.run_macro(excel, workbook.Name, macro_name)
                        break
                    except Exception as error:
                        last_error = error
                else:
                    self.emit(
                        "excel_macro_warning",
                        itemId=item.get("id"),
                        macroCandidates=MACRO_TOGGLE_SECTIONS_CANDIDATES,
                        message=f"Bereichsmakro nicht verfügbar: {last_error}",
                    )
            try:
                self.run_macro(excel, workbook.Name, MACRO_HIDE_EMPTY_ROWS)
            except Exception as error:
                self.emit(
                    "excel_macro_warning",
                    itemId=item.get("id"),
                    macro=MACRO_HIDE_EMPTY_ROWS,
                    message=f"Makro fehlgeschlagen: {error}",
                )
            self.check_cancelled()

            if customer_assignment_failed_without_template:
                project_title = self.fallback_title_from_inspector_and_date(
                    inspector_for_fallback,
                    exam_date_for_fallback,
                    excel_path,
                )
            else:
                project_title = self.resolve_project_title(workbook, excel_path)

            output_path = self.build_output_excel_path(excel_path, project_title)
            try:
                if output_path == excel_path:
                    workbook.Save()
                else:
                    workbook.SaveAs(str(output_path.resolve()), FileFormat=workbook.FileFormat)
            except Exception as error:
                self.emit(
                    "excel_save_warning",
                    itemId=item.get("id"),
                    excelPath=str(excel_path),
                    message=f"SaveAs fehlgeschlagen, versuche Save auf Originaldatei: {error}",
                )
                workbook.Save()
                output_path = excel_path

            workbook.Close(SaveChanges=True)
            workbook = None
        finally:
            if workbook is not None:
                workbook.Close(SaveChanges=False)
            if excel is not None:
                excel.Quit()
        return output_path

    def run(self) -> int:
        items = self.job.get("items", [])
        self.run_started_monotonic = time.monotonic()
        self.emit("run_started", itemCount=len(items))
        
        # Sicherheit: Mashup zu Anfang beenden (falls noch aktiv)
        self.terminate_mashup_loader()
        
        succeeded_count = 0
        failed_count = 0
        skipped_count = 0
        for index, item in enumerate(items, start=1):
            item_id = item.get("id", str(index))
            current_excel_path: str | None = None
            current_occ_path: str | None = None
            try:
                self.check_cancelled()
                if not item.get("enabled", True):
                    self.emit("item_skipped", itemId=item_id, reason="Ordner wurde vor Verarbeitung deaktiviert")
                    skipped_count += 1
                    self.skipped.append({"itemId": str(item_id), "reason": "Ordner wurde vor Verarbeitung deaktiviert"})
                    continue
                if item.get("mappingStatus") != "eindeutig":
                    self.emit("item_skipped", itemId=item_id, reason="Zuordnung vor Start nicht eindeutig")
                    skipped_count += 1
                    self.skipped.append({"itemId": str(item_id), "reason": "Zuordnung vor Start nicht eindeutig"})
                    continue
                excel_groups = self.resolve_excel_groups(item)
                if not excel_groups:
                    self.emit("item_skipped", itemId=item_id, reason="Keine OCC-/Excel-Zuordnung vorhanden")
                    skipped_count += 1
                    self.skipped.append({"itemId": str(item_id), "reason": "Keine OCC-/Excel-Zuordnung vorhanden"})
                    continue
                self.emit("item_started", itemId=item_id, index=index, total=len(items))

                # Gewünschte Reihenfolge pro Ordner:
                # 1. Mashup beenden vor OCC-Export
                # 2. Alle OCCs der Excel-Gruppe exportieren
                # 3. Excel öffnen, Daten importieren, Makros ausführen
                # 4. Mashup beenden nach Excel-Import (vor nächster Gruppe)
                for excel_path, occ_paths in excel_groups:
                    working_excel_path = excel_path
                    current_excel_path = str(working_excel_path)
                    if not occ_paths or not working_excel_path.is_file():
                        raise FileNotFoundError(f"OCC- oder Excel-Datei fehlt für Zuordnung: {working_excel_path}")

                    self.terminate_mashup_loader()
                    for occ_path in occ_paths:
                        current_occ_path = str(occ_path)
                        self.emit("occ_started", itemId=item_id, occPath=str(occ_path), excelPath=str(working_excel_path))
                        self.export_occ(occ_path)
                        self.emit("occ_completed", itemId=item_id, occPath=str(occ_path), excelPath=str(working_excel_path))

                    self.emit("excel_started", itemId=item_id, excelPath=str(working_excel_path))
                    working_excel_path = self.refresh_excel(working_excel_path, item)
                    current_excel_path = str(working_excel_path)
                    self.emit("excel_completed", itemId=item_id, excelPath=str(working_excel_path))
                    
                    self.terminate_mashup_loader()

                archived = self.archive_processed_folder(item)
                self.emit("item_completed", itemId=item_id, archived=archived)
                succeeded_count += 1
            except CancellationRequested:
                self.emit("run_cancelled", itemId=item_id, elapsedSeconds=self.elapsed_seconds())
                return 2
            except Exception as error:
                self.emit(
                    "item_failed",
                    itemId=item_id,
                    message=str(error),
                    excelPath=current_excel_path,
                    occPath=current_occ_path,
                )
                failed_count += 1
                failure_entry: dict[str, str] = {"itemId": str(item_id), "message": str(error)}
                if current_excel_path:
                    failure_entry["excelPath"] = current_excel_path
                if current_occ_path:
                    failure_entry["occPath"] = current_occ_path
                self.failures.append(failure_entry)

        report_path = None
        try:
            if failed_count > 0 or skipped_count > 0 or self.archive_warnings:
                report_path = self.write_report_if_needed(succeeded_count, failed_count, skipped_count)
                if report_path:
                    self.emit("run_report_written", reportPath=report_path)
        except Exception as error:
            self.emit("run_report_failed", message=str(error))

        self.emit(
            "run_completed",
            elapsedSeconds=self.elapsed_seconds(),
            succeededCount=succeeded_count,
            failedCount=failed_count,
            skippedCount=skipped_count,
            archiveWarningCount=len(self.archive_warnings),
            reportPath=report_path,
        )
        return 0 if failed_count == 0 else 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Unbeaufsichtigter Omicron-OCC-Worker")
    parser.add_argument("job", type=Path, help="JSON-Datei mit vorab bestätigten Verarbeitungseinheiten")
    parser.add_argument("--cancel-file", type=Path, help="Vorhandene Datei löst kontrollierten Abbruch aus")
    args = parser.parse_args()
    with args.job.open(encoding="utf-8") as job_file:
        job = json.load(job_file)
    return Worker(job, args.cancel_file).run()


if __name__ == "__main__":
    sys.exit(main())