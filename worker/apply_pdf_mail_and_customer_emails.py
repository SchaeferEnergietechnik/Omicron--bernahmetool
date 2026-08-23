from __future__ import annotations

import argparse
import os
import shutil
import sys
import tempfile
import time
import xml.etree.ElementTree as ET
from zipfile import ZipFile
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_TARGETS_REAL = [
    Path("samples/V20d_Übergeordneter_Entkupplungsschutz.xlsm"),
    Path("samples/topics/excel-basis/V20d_Übergeordneter_Entkupplungsschutz.xlsm"),
    Path("samples/topics/excel-basis/V20f_Schutzprüfprotokoll-Checkliste.xlsm"),
]

DEFAULT_TARGETS_FALLBACK = [
    Path("samples/V20d_Uebergeordneter_Entkupplungsschutz.xlsm"),
    Path("samples/topics/excel-basis/V20d_Uebergeordneter_Entkupplungsschutz.xlsm"),
    Path("samples/topics/excel-basis/V20f_Schutzpruefprotokoll-Checkliste.xlsm"),
    Path("V20d_Übergeordneter_Entkupplungsschutz.xlsm"),
    Path("V20d_Uebergeordneter_Entkupplungsschutz.xlsm"),
]

DEFAULT_SOURCE = Path("samples/topics/excel-basis/Muster_Termine 17.08.2026.xlsx")
CUSTOMER_SHEET = "Kunden"
SOURCE_CANDIDATE_SHEETS = ("Kundenadressen", "Kunden")
LOGO_REFERENCE_CANDIDATES = [
    Path("samples/V19m_Übergeordneter_Entkupplungsschutz.xlsm"),
    Path("samples/V19g_Übergeordneter_Entkupplungsschutz.xlsm"),
    Path("samples/topics/V19m_Übergeordneter_Entkupplungsschutz.xlsm"),
    Path("samples/topics/excel-basis/V19m_Übergeordneter_Entkupplungsschutz.xlsm"),
]

# Known edge ranges where validation transfer can fail due to merges/protection quirks.
CHECKLIST_VALIDATION_EDGE_RANGES = (
    "$A$14:$D$17",
    "$B$9:$D$9",
    "$B$19:$D$19",
    "$D$65:$D$66",
)

FORM_CODE = r'''Option Explicit
Private Const EMAIL_CHECKBOX_NAME As String = "chkDirektEmail"

Private Sub UserForm_Initialize()
    EnsureEmailCheckbox
End Sub

Private Sub EnsureEmailCheckbox()
    Dim ctrl As Object
    On Error Resume Next
    Set ctrl = Me.Controls(EMAIL_CHECKBOX_NAME)
    On Error GoTo 0

    If ctrl Is Nothing Then
        Set ctrl = Me.Controls.Add("Forms.CheckBox.1", EMAIL_CHECKBOX_NAME, True)
        With ctrl
            .Caption = "PDF direkt per E-Mail vorbereiten"
            .Left = 18
            .Top = 166
            .Width = 220
            .Height = 18
            .Value = False
        End With
    End If
End Sub

Private Sub cmdOK_Click()
    Dim wb As Workbook
    Set wb = ThisWorkbook

    Dim pfad As String
    Dim dateiname As String
    Dim vollDateiname As String
    Dim vollPfad As String
    Dim schutz As String
    Dim stationPrefix As String
    Dim wsProt As Worksheet
    Dim exportiertePdfs As Collection
    Set exportiertePdfs = New Collection

    pfad = wb.Path & "\\"

    If optEntkupplung.Value = True Then
        schutz = "Entkupplungsschutz"
    ElseIf optUMZ.Value = True Then
        schutz = "UMZ-Schutz"
    Else
        MsgBox "Bitte eine Schutzart auswaehlen!", vbExclamation
        Exit Sub
    End If

    Set wsProt = wb.Sheets("Prüfprotokoll")
    If IstStationGesperrt(wsProt) Then
        stationPrefix = "Station-gesperrt_"
    Else
        stationPrefix = ""
    End If

    If chkBlatt1.Value = True Then
        dateiname = stationPrefix & wb.Sheets("Allgemeine Angaben").Range("C5").Value & "_" & _
                    wb.Sheets("Prüfprotokoll").Range("H13").Value & "_" & _
                    IIf(schutz = "Entkupplungsschutz", "Übergeordneter-Schutz", "UMZ-Schutz")

        vollDateiname = DateiMitVersion(pfad, dateiname)
        vollPfad = pfad & vollDateiname

        wb.Sheets("Prüfprotokoll").ExportAsFixedFormat Type:=xlTypePDF, _
            Filename:=vollPfad, Quality:=xlQualityStandard, _
            IncludeDocProperties:=True, IgnorePrintAreas:=False, OpenAfterPublish:=False

        exportiertePdfs.Add vollPfad
    End If

    If chkBlatt2.Value = True Then
        dateiname = stationPrefix & wb.Sheets("Allgemeine Angaben").Range("C5").Value & "_" & _
                    wb.Sheets("Prüfprotokoll").Range("H13").Value & "_Wandlerprüfprotokoll"

        vollDateiname = DateiMitVersion(pfad, dateiname)
        vollPfad = pfad & vollDateiname

        wb.Sheets("Wandlerprüfprotokoll").ExportAsFixedFormat Type:=xlTypePDF, _
            Filename:=vollPfad, Quality:=xlQualityStandard, _
            IncludeDocProperties:=True, IgnorePrintAreas:=False, OpenAfterPublish:=False

        exportiertePdfs.Add vollPfad
    End If

    If chkBlatt3.Value = True Then
        dateiname = stationPrefix & wb.Sheets("Allgemeine Angaben").Range("C6").Value & "_" & _
                    wb.Sheets("Prüfprotokoll").Range("H13").Value & "_Untergeordneter-Schutz"

        vollDateiname = DateiMitVersion(pfad, dateiname)
        vollPfad = pfad & vollDateiname

        wb.Sheets("Protokolluntergeord. Schutz EZE").ExportAsFixedFormat Type:=xlTypePDF, _
            Filename:=vollPfad, Quality:=xlQualityStandard, _
            IncludeDocProperties:=True, IgnorePrintAreas:=False, OpenAfterPublish:=False

        exportiertePdfs.Add vollPfad
    End If

    If exportiertePdfs.Count = 0 Then
        MsgBox "Es wurde kein Blatt fuer den PDF-Export ausgewaehlt.", vbExclamation
        Exit Sub
    End If

    MsgBox "PDF(s) erfolgreich erstellt im Ordner: " & pfad

    If IsEmailCheckboxSelected() Then
        SendePdfsPerOutlook exportiertePdfs
    End If

    Unload Me
End Sub

Private Function IsEmailCheckboxSelected() As Boolean
    Dim ctrl As Object

    On Error Resume Next
    Set ctrl = Me.Controls(EMAIL_CHECKBOX_NAME)
    On Error GoTo 0

    If ctrl Is Nothing Then
        IsEmailCheckboxSelected = False
    Else
        IsEmailCheckboxSelected = CBool(ctrl.Value)
    End If
End Function

Private Sub SendePdfsPerOutlook(ByVal exportiertePdfs As Collection)
    Dim wb As Workbook
    Dim wsAngaben As Worksheet
    Dim wsKunden As Worksheet
    Dim wsCheck As Worksheet
    Dim kundeText As String
    Dim emailTo As String
    Dim bemerkungen As String
    Dim betreff As String
    Dim projektname As String
    Dim protokollLabel As String
    Dim bodyText As String
    Dim htmlBody As String
    Dim logoPath As String

    Set wb = ThisWorkbook
    Set wsAngaben = wb.Sheets("Allgemeine Angaben")
    Set wsKunden = wb.Sheets("Kunden")
    Set wsCheck = wb.Sheets("Schutzprüf-Checkliste")

    kundeText = CStr(wsAngaben.Range("C2").Value)
    emailTo = HoleEmailAusKundenblatt(wsKunden, kundeText)

    If Trim$(emailTo) = "" Then
        MsgBox "Keine E-Mail-Adresse fuer den Kunden gefunden. Bitte im Blatt 'Kunden' in Spalte B pflegen.", vbExclamation
        Exit Sub
    End If

    bemerkungen = HoleBemerkungenAusCheckliste(wsCheck)

    projektname = Trim$(CStr(wsCheck.Range("B3").Value))
    If projektname = "" Then
        projektname = Trim$(CStr(wsAngaben.Range("C5").Value))
    End If
    protokollLabel = IIf(exportiertePdfs.Count = 1, "Schutzprüfprotokoll", "Schutzprüfprotokolle")
    betreff = protokollLabel & IIf(projektname <> "", " - " & projektname, "")
    bodyText = ErzeugeStandardMailtext(kundeText, bemerkungen, projektname, exportiertePdfs.Count)
    logoPath = FindeLogoPfad(wb.Path)
    htmlBody = ErzeugeHtmlMailtext(kundeText, bemerkungen, logoPath, projektname, exportiertePdfs.Count)

    Dim olApp As Object
    Dim olMail As Object
    Dim i As Long

    On Error GoTo outlook_fehler
    Set olApp = CreateObject("Outlook.Application")
    Set olMail = olApp.CreateItem(0)

    With olMail
        .To = emailTo
        .Subject = betreff
        .Body = bodyText

        For i = 1 To exportiertePdfs.Count
            If Len(Dir(CStr(exportiertePdfs(i)))) > 0 Then
                .Attachments.Add CStr(exportiertePdfs(i))
            End If
        Next i

        If logoPath <> "" Then
            Dim logoAttachment As Object
            Set logoAttachment = .Attachments.Add(logoPath)
            logoAttachment.PropertyAccessor.SetProperty _
                "http://schemas.microsoft.com/mapi/proptag/0x3712001F", "geslogo"
        End If

        .HTMLBody = htmlBody

        .Display
    End With

    MsgBox "Outlook-E-Mail wurde vorbereitet.", vbInformation
    Exit Sub

outlook_fehler:
    MsgBox "Outlook konnte nicht gestartet werden oder der Entwurf konnte nicht erstellt werden.", vbCritical
End Sub

Private Function ErzeugeStandardMailtext(ByVal kundeText As String, ByVal bemerkungen As String, ByVal projektname As String, ByVal anzahlPdfs As Long) As String
    Dim text As String
    Dim protokollLabel As String
    Dim projektTeil As String

    protokollLabel = IIf(anzahlPdfs = 1, "Schutzprüfprotokoll", "Schutzprüfprotokolle")
    If Trim$(projektname) <> "" Then
        projektTeil = " zum Projekt """ & projektname & """"
    Else
        projektTeil = ""
    End If

    text = "Sehr geehrte Damen und Herren," & vbCrLf & vbCrLf & _
        "anbei erhalten Sie " & IIf(anzahlPdfs = 1, "das aktuelle ", "die aktuellen ") & protokollLabel & projektTeil & " als PDF-Anhang." & vbCrLf & _
        "Bitte prüfen Sie die Unterlagen und melden Sie sich gerne bei Rückfragen." & vbCrLf & vbCrLf

    If Trim$(bemerkungen) <> "" Then
        text = text & "Bemerkungen aus der Checkliste:" & vbCrLf & bemerkungen & vbCrLf & vbCrLf
    End If

    text = text & "Herzlichen Dank." & vbCrLf & vbCrLf & _
        "Mit freundlichen Grüßen" & vbCrLf & vbCrLf & _
        "Ihr Team von G.E.S. Energietechnik GmbH" & vbCrLf & vbCrLf & _
        "Der Inhalt dieser Email ist ausschließlich für den bezeichneten Adressaten bestimmt." & vbCrLf & _
        "Falls Sie nicht der vorgesehene Adressat dieser Email oder dessen Vertreter sein sollten," & vbCrLf & _
        "so beachten Sie bitte, dass jede Form der Kenntnisnahme, Veröffentlichung," & vbCrLf & _
        "Vervielfältigung oder Weitergabe des Inhalts dieser Email unzulässig ist." & vbCrLf & _
        "Wir bitten Sie, sich in diesem Fall mit dem Absender der Email in Verbindung zu setzen." & vbCrLf & vbCrLf & _
        "G.E.S. Energietechnik GmbH: Sitz der Gesellschaft: Altmärkische Wische - Amtsgericht Stendal," & vbCrLf & _
        "HRB 30020 - Geschäftsführer: Gunnar Schäfer"

    ErzeugeStandardMailtext = text
End Function

Private Function ErzeugeHtmlMailtext(ByVal kundeText As String, ByVal bemerkungen As String, ByVal logoPath As String, ByVal projektname As String, ByVal anzahlPdfs As Long) As String
    Dim text As String
    Dim bemerkHtml As String
    Dim protokollLabel As String
    Dim projektTeil As String

    bemerkHtml = Replace(HTMLEncode(bemerkungen), vbCrLf, "<br>")
    protokollLabel = IIf(anzahlPdfs = 1, "Schutzprüfprotokoll", "Schutzprüfprotokolle")
    If Trim$(projektname) <> "" Then
        projektTeil = " zum Projekt &bdquo;" & HTMLEncode(projektname) & "&ldquo;"
    Else
        projektTeil = ""
    End If

    text = "<html><body style='font-family:Calibri,Arial,sans-serif;font-size:11pt;'>" & _
           "<p>Sehr geehrte Damen und Herren,</p>" & _
             "<p>anbei erhalten Sie " & IIf(anzahlPdfs = 1, "das aktuelle ", "die aktuellen ") & HTMLEncode(protokollLabel) & projektTeil & " als PDF-Anhang.<br>" & _
            "Bitte prüfen Sie die Unterlagen und melden Sie sich gerne bei Rückfragen.</p>"

    If Trim$(bemerkungen) <> "" Then
        text = text & "<p><b>Bemerkungen aus der Checkliste:</b><br>" & bemerkHtml & "</p>"
    End If

    text = text & "<p>Herzlichen Dank.</p>" & _
                  "<p>Mit freundlichen Grüßen</p>" & _
                  "<p>Ihr Team von G.E.S. Energietechnik GmbH</p>" & _
                  "<p style='font-size:9pt;color:#555;'>" & _
                  "Der Inhalt dieser Email ist ausschließlich für den bezeichneten Adressaten bestimmt.<br>" & _
                  "Falls Sie nicht der vorgesehene Adressat dieser Email oder dessen Vertreter sein sollten,<br>" & _
                  "so beachten Sie bitte, dass jede Form der Kenntnisnahme, Veröffentlichung,<br>" & _
                  "Vervielfältigung oder Weitergabe des Inhalts dieser Email unzulässig ist.<br>" & _
                  "Wir bitten Sie, sich in diesem Fall mit dem Absender der Email in Verbindung zu setzen.<br><br>" & _
                  "G.E.S. Energietechnik GmbH: Sitz der Gesellschaft: Altmärkische Wische - Amtsgericht Stendal,<br>" & _
                  "HRB 30020 - Geschäftsführer: Gunnar Schäfer" & _
                  "</p>"

    If logoPath <> "" Then
        text = text & "<p><img src='cid:geslogo' style='max-width:220px;height:auto;'></p>"
    End If

    text = text & "</body></html>"
    ErzeugeHtmlMailtext = text
End Function

Private Function HTMLEncode(ByVal text As String) As String
    Dim t As String
    t = text
    t = Replace(t, "&", "&amp;")
    t = Replace(t, "<", "&lt;")
    t = Replace(t, ">", "&gt;")
    t = Replace(t, Chr(34), "&quot;")
    HTMLEncode = t
End Function

Private Function IstStationGesperrt(ByVal wsProt As Worksheet) As Boolean
    Dim zeile As Long
    Dim statusWert As String

    For zeile = 167 To 175
        statusWert = Trim$(CStr(wsProt.Cells(zeile, "J").Value))
        If StrComp(statusWert, "NICHT OK", vbTextCompare) = 0 Then
            IstStationGesperrt = True
            Exit Function
        End If
    Next zeile

    IstStationGesperrt = False
End Function

Private Function FindeLogoPfad(ByVal basisPfad As String) As String
    Dim kandidaten(1 To 7) As String
    Dim i As Long

    kandidaten(1) = basisPfad & "\\logo.png"
    kandidaten(2) = basisPfad & "\\Logo.png"
    kandidaten(3) = basisPfad & "\\logo.jpg"
    kandidaten(4) = basisPfad & "\\Logo.jpg"
    kandidaten(5) = basisPfad & "\\logo.jpeg"
    kandidaten(6) = basisPfad & "\\ges-logo.png"
    kandidaten(7) = basisPfad & "\\GES_Logo.png"

    For i = 1 To 7
        If Len(Dir(kandidaten(i))) > 0 Then
            FindeLogoPfad = kandidaten(i)
            Exit Function
        End If
    Next i
End Function

Private Function HoleBemerkungenAusCheckliste(ByVal ws As Worksheet) As String
    Dim zeile As Long
    Dim eintrag As String
    Dim gesammelt As String

    ' Unteres Bemerkungsfeld in V19m/V20a: B138:B147
    For zeile = 138 To 147
        eintrag = Trim$(CStr(ws.Cells(zeile, "B").Value))
        If eintrag <> "" Then
            If gesammelt <> "" Then
                gesammelt = gesammelt & vbCrLf
            End If
            gesammelt = gesammelt & CStr(zeile - 137) & ". " & eintrag
        End If
    Next zeile

    HoleBemerkungenAusCheckliste = gesammelt
End Function

Private Function HoleEmailAusKundenblatt(ByVal ws As Worksheet, ByVal kundenwert As String) As String
    Dim suchKey As String
    Dim zeile As Long
    Dim kandidat As String
    Dim kandidatKey As String
    Dim email As String

    suchKey = NormalisiereKundenname(kundenwert)
    If suchKey = "" Then
        Exit Function
    End If

    For zeile = 1 To ws.Cells(ws.Rows.Count, "A").End(xlUp).Row
        kandidat = CStr(ws.Cells(zeile, "A").Value)
        kandidatKey = NormalisiereKundenname(kandidat)
        email = Trim$(CStr(ws.Cells(zeile, "B").Value))

        If email <> "" Then
            If kandidatKey = suchKey Then
                HoleEmailAusKundenblatt = email
                Exit Function
            End If
        End If
    Next zeile

    For zeile = 1 To ws.Cells(ws.Rows.Count, "A").End(xlUp).Row
        kandidat = CStr(ws.Cells(zeile, "A").Value)
        kandidatKey = NormalisiereKundenname(kandidat)
        email = Trim$(CStr(ws.Cells(zeile, "B").Value))

        If email <> "" Then
            If (InStr(1, kandidatKey, suchKey, vbTextCompare) > 0) Or _
               (InStr(1, suchKey, kandidatKey, vbTextCompare) > 0) Then
                HoleEmailAusKundenblatt = email
                Exit Function
            End If
        End If
    Next zeile
End Function

Private Function NormalisiereKundenname(ByVal rawText As String) As String
    Dim t As String
    Dim teile() As String

    t = Replace(rawText, vbCr, vbLf)
    teile = Split(t, vbLf)
    t = Trim$(teile(0))

    t = LCase$(t)
    t = Replace(t, "ae", "a")
    t = Replace(t, "oe", "o")
    t = Replace(t, "ue", "u")
    t = Replace(t, "ss", "s")

    t = Replace(t, "gmbh", "")
    t = Replace(t, "co. kg", "")
    t = Replace(t, "co kg", "")
    t = Replace(t, "kg", "")
    t = Replace(t, "ag", "")

    t = Replace(t, "-", " ")
    t = Replace(t, ",", " ")
    t = Replace(t, ".", " ")

    Do While InStr(t, "  ") > 0
        t = Replace(t, "  ", " ")
    Loop

    NormalisiereKundenname = Trim$(t)
End Function

Function DateiMitVersion(pfad As String, BasisName As String) As String
    Dim version As Long
    Dim vollName As String

    version = 1
    vollName = BasisName & "_v" & version & ".pdf"

    Do While Dir(pfad & vollName) <> ""
        version = version + 1
        vollName = BasisName & "_v" & version & ".pdf"
    Loop

    DateiMitVersion = vollName
End Function

Private Sub txtVersion_Change()

End Sub
'''

DATE_MACRO_CODE = r'''Option Explicit

Public Sub AktuellesDatum()
    ApplyCurrentDateToSelection
End Sub

Public Sub Datum()
    ApplyCurrentDateToSelection
End Sub

Private Sub ApplyCurrentDateToSelection()
    On Error Resume Next

    If TypeName(Selection) = "Range" Then
        Selection.Value = Date
        Selection.NumberFormat = "dd.mm.yyyy"
    End If
End Sub
'''


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Erweitert den PDF-Export in frmPDFDruck um Outlook-E-Mail-Versand "
            "und pflegt E-Mail-Adressen in Kunden!B anhand Muster_Termine."
        )
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=DEFAULT_SOURCE,
        help="Quelle fuer Kunde->E-Mail (xlsx, Blatt Kundenadressen/Kunden)",
    )
    parser.add_argument(
        "--targets",
        type=Path,
        nargs="+",
        default=DEFAULT_TARGETS_REAL,
        help="Ziel-xlsm-Dateien, die angepasst werden sollen",
    )
    parser.add_argument(
        "--backup-suffix",
        default=".bak_before_pdf_mail",
        help="Suffix fuer Sicherung vor der Endung .xlsm",
    )
    parser.add_argument(
        "--visible",
        action="store_true",
        help="Excel sichtbar starten (Debug)",
    )
    return parser.parse_args()


def ensure_exists(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Datei nicht gefunden: {path}")


def _unique_paths(paths: list[Path]) -> list[Path]:
    seen: set[str] = set()
    result: list[Path] = []
    for path in paths:
        key = str(path).lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(path)
    return result


def resolve_source_path(source_path: Path) -> Path:
    if source_path.exists():
        return source_path

    fallback_candidates: list[Path] = [
        Path("samples/topics/excel-basis/Muster_Termine 17.08.2026.xlsx"),
        Path("samples/Muster_Termine 17.08.2026.xlsx"),
        Path("Muster_Termine 17.08.2026.xlsx"),
    ]
    for candidate in fallback_candidates:
        if candidate.exists():
            return candidate

    recursive_hits = sorted(Path(".").glob("**/Muster_Termine*.xlsx"))
    if recursive_hits:
        return recursive_hits[0]

    termine_hits = sorted(Path(".").glob("**/Termine*.xlsx"))
    if termine_hits:
        return termine_hits[0]

    raise FileNotFoundError(
        f"Datei nicht gefunden: {source_path}. "
        "Bitte --source explizit angeben."
    )


def resolve_target_paths(targets: list[Path]) -> list[Path]:
    existing = [target for target in targets if target.exists()]
    if existing:
        return _unique_paths(existing)

    candidates: list[Path] = []
    for candidate in DEFAULT_TARGETS_REAL + DEFAULT_TARGETS_FALLBACK:
        if candidate.exists():
            candidates.append(candidate)

    if candidates:
        return _unique_paths(candidates)

    recursive_hits: list[Path] = []
    for pattern in [
        "**/*V20d*Übergeordneter_Entkupplungsschutz*.xlsm",
        "**/*V20D*Übergeordneter_Entkupplungsschutz*.xlsm",
        "**/*v20d*Übergeordneter_Entkupplungsschutz*.xlsm",
        "**/*V20d*Uebergeordneter_Entkupplungsschutz*.xlsm",
        "**/*V20D*Uebergeordneter_Entkupplungsschutz*.xlsm",
        "**/*v20d*Uebergeordneter_Entkupplungsschutz*.xlsm",
        "**/*V20*Schutzprüfprotokoll-Checkliste*.xlsm",
        "**/*V20*Schutzpruefprotokoll-Checkliste*.xlsm",
        "**/*v20*Schutzprüfprotokoll-Checkliste*.xlsm",
        "**/*v20*Schutzpruefprotokoll-Checkliste*.xlsm",
    ]:
        recursive_hits.extend(sorted(Path(".").glob(pattern)))

    filtered = [
        path
        for path in recursive_hits
        if "bak" not in path.name.lower() and not path.name.startswith("~$")
    ]
    if filtered:
        return _unique_paths(filtered)

    raise FileNotFoundError(
        "Keine passende V20d-Ziel-XLSM gefunden. Bitte mit --targets die V20d-Datei explizit angeben."
    )


def resolve_logo_reference_path(target_path: Path) -> Path | None:
    for candidate in LOGO_REFERENCE_CANDIDATES:
        if candidate.exists() and candidate.resolve() != target_path.resolve():
            return candidate

    recursive_hits = sorted(Path(".").glob("**/*V19*Übergeordneter_Entkupplungsschutz*.xlsm"))
    for hit in recursive_hits:
        if hit.resolve() != target_path.resolve():
            return hit

    recursive_hits_ascii = sorted(Path(".").glob("**/*V19*Uebergeordneter_Entkupplungsschutz*.xlsm"))
    for hit in recursive_hits_ascii:
        if hit.resolve() != target_path.resolve():
            return hit

    return None


def _shape_intersects_d7_e7(shape) -> bool:
    try:
        tl_row = int(shape.TopLeftCell.Row)
        tl_col = int(shape.TopLeftCell.Column)
        br_row = int(shape.BottomRightCell.Row)
        br_col = int(shape.BottomRightCell.Column)
    except Exception:
        return False

    return not (br_row < 7 or tl_row > 7 or br_col < 4 or tl_col > 5)


def restore_logos_from_reference(excel_app, target_workbook, reference_path: Path) -> int:
    ref_workbook = None
    restored = 0
    try:
        ref_workbook = excel_app.Workbooks.Open(str(reference_path.resolve()), ReadOnly=True, AddToMru=False)
        ws_src = ref_workbook.Worksheets("Allgemeine Angaben")
        ws_dst = target_workbook.Worksheets("Allgemeine Angaben")

        # Restore possible formula/text content in D7:E7.
        ws_dst.Range("D7:E7").Formula = ws_src.Range("D7:E7").Formula

        # Remove existing shapes in D7:E7 before re-pasting from V19 reference.
        for i in range(int(ws_dst.Shapes.Count), 0, -1):
            shp = ws_dst.Shapes(i)
            if _shape_intersects_d7_e7(shp):
                try:
                    shp.Delete()
                except Exception:
                    pass

        for i in range(1, int(ws_src.Shapes.Count) + 1):
            shp = ws_src.Shapes(i)
            if not _shape_intersects_d7_e7(shp):
                continue

            left = float(shp.Left)
            top = float(shp.Top)
            width = float(shp.Width)
            height = float(shp.Height)

            shp.Copy()
            ws_dst.Paste()
            pasted = ws_dst.Shapes(int(ws_dst.Shapes.Count))
            pasted.Left = left
            pasted.Top = top
            pasted.Width = width
            pasted.Height = height
            restored += 1

        return restored
    finally:
        if ref_workbook is not None:
            ref_workbook.Close(SaveChanges=False)


def restore_checkliste_dropdowns_from_reference(excel_app, target_workbook, reference_path: Path) -> int:
    ref_workbook = None
    try:
        ref_workbook = excel_app.Workbooks.Open(str(reference_path.resolve()), ReadOnly=True, AddToMru=False)
        ws_src = ref_workbook.Worksheets("Schutzprüf-Checkliste")
        ws_dst = target_workbook.Worksheets("Schutzprüf-Checkliste")

        try:
            ws_dst.Unprotect()
        except Exception:
            pass

        transferred_areas = 0
        recovered_areas = 0
        failed_areas: list[str] = []

        def _copy_validation(src_range, dst_range) -> bool:
            try:
                src_validation = src_range.Validation
                validation_type = int(src_validation.Type)
                # xlValidateInputOnly = 0 (no dropdown rule to recreate)
                if validation_type == 0:
                    return False

                alert_style = int(src_validation.AlertStyle)
                operator = int(src_validation.Operator)
                formula1 = src_validation.Formula1
                formula2 = src_validation.Formula2

                try:
                    dst_range.Validation.Delete()
                except Exception:
                    pass

                dst_range.Validation.Add(
                    Type=validation_type,
                    AlertStyle=alert_style,
                    Operator=operator,
                    Formula1=formula1,
                    Formula2=formula2,
                )

                dst_validation = dst_range.Validation
                dst_validation.IgnoreBlank = src_validation.IgnoreBlank
                dst_validation.InCellDropdown = src_validation.InCellDropdown
                dst_validation.InputTitle = src_validation.InputTitle
                dst_validation.ErrorTitle = src_validation.ErrorTitle
                dst_validation.InputMessage = src_validation.InputMessage
                dst_validation.ErrorMessage = src_validation.ErrorMessage
                dst_validation.ShowInput = src_validation.ShowInput
                dst_validation.ShowError = src_validation.ShowError

                return int(dst_range.Validation.Type) != 0
            except Exception:
                return False

        try:
            # xlCellTypeAllValidation = -4174
            validated_cells = ws_src.Cells.SpecialCells(-4174)
            for i in range(1, int(validated_cells.Areas.Count) + 1):
                area = validated_cells.Areas(i)
                address = str(area.Address)
                dst_range = ws_dst.Range(address)

                try:
                    dst_range.Validation.Delete()
                except Exception:
                    pass

                try:
                    # First try direct validation paste; this preserves complex source rules best.
                    area.Copy()
                    dst_range.PasteSpecial(Paste=6)  # xlPasteValidation
                    excel_app.CutCopyMode = False

                    if int(dst_range.Validation.Type) != 0:
                        transferred_areas += 1
                        continue
                except Exception:
                    pass

                if _copy_validation(area, dst_range):
                    transferred_areas += 1
                else:
                    failed_areas.append(address)
        except Exception:
            # Fallback for workbooks where SpecialCells is unavailable/unreliable.
            ws_src.UsedRange.Copy()
            ws_dst.UsedRange.PasteSpecial(Paste=6)
            transferred_areas = 1

        if failed_areas:
            still_failed: list[str] = []
            for address in failed_areas:
                src_area = ws_src.Range(address)
                area_recovered = False

                for i in range(1, int(src_area.Cells.Count) + 1):
                    src_cell = src_area.Cells(i)
                    dst_cell = ws_dst.Cells(src_cell.Row, src_cell.Column)

                    try:
                        if bool(src_cell.MergeCells):
                            src_anchor = src_cell.MergeArea.Cells(1, 1)
                            if str(src_cell.Address) != str(src_anchor.Address):
                                continue
                            src_target = src_cell.MergeArea
                        else:
                            src_target = src_cell
                    except Exception:
                        src_target = src_cell

                    try:
                        if bool(dst_cell.MergeCells):
                            dst_anchor = dst_cell.MergeArea.Cells(1, 1)
                            if str(dst_cell.Address) != str(dst_anchor.Address):
                                continue
                            dst_target = dst_cell.MergeArea
                        else:
                            dst_target = dst_cell
                    except Exception:
                        dst_target = dst_cell

                    if _copy_validation(src_target, dst_target):
                        area_recovered = True

                if area_recovered:
                    recovered_areas += 1
                else:
                    still_failed.append(address)

            failed_areas = still_failed

        # Deterministic final pass for known edge ranges.
        forced_recovered = 0
        for address in CHECKLIST_VALIDATION_EDGE_RANGES:
            try:
                src_area = ws_src.Range(address)
                dst_area = ws_dst.Range(address)
                if _copy_validation(src_area, dst_area):
                    forced_recovered += 1
                    continue

                area_recovered = False
                for i in range(1, int(src_area.Cells.Count) + 1):
                    src_cell = src_area.Cells(i)
                    dst_cell = ws_dst.Cells(src_cell.Row, src_cell.Column)

                    try:
                        if bool(src_cell.MergeCells):
                            src_anchor = src_cell.MergeArea.Cells(1, 1)
                            if str(src_cell.Address) != str(src_anchor.Address):
                                continue
                            src_target = src_cell.MergeArea
                        else:
                            src_target = src_cell
                    except Exception:
                        src_target = src_cell

                    try:
                        if bool(dst_cell.MergeCells):
                            dst_anchor = dst_cell.MergeArea.Cells(1, 1)
                            if str(dst_cell.Address) != str(dst_anchor.Address):
                                continue
                            dst_target = dst_cell.MergeArea
                        else:
                            dst_target = dst_cell
                    except Exception:
                        dst_target = dst_cell

                    if _copy_validation(src_target, dst_target):
                        area_recovered = True

                if area_recovered:
                    forced_recovered += 1
            except Exception:
                continue

        excel_app.CutCopyMode = False
        if failed_areas:
            print(f"Hinweis: {len(failed_areas)} Validierungs-Bereiche konnten nicht 1:1 kopiert werden")
            print(f"Hinweis: Betroffene Bereiche: {', '.join(failed_areas[:8])}")
        if recovered_areas > 0:
            print(f"Hinweis: Zusätzliche Validierungs-Bereiche zellweise repariert: {recovered_areas}")
        if forced_recovered > 0:
            print(f"Hinweis: Edge-Validierungs-Bereiche gezielt nachgezogen: {forced_recovered}")

        return transferred_areas + recovered_areas + forced_recovered
    finally:
        try:
            excel_app.CutCopyMode = False
        except Exception:
            pass
        if ref_workbook is not None:
            ref_workbook.Close(SaveChanges=False)


def _extract_x14_dropdown_validations(workbook_path: Path, sheet_name: str) -> list[tuple[str, str]]:
    ns_main = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    ns_rel = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    ns_x14 = "{http://schemas.microsoft.com/office/spreadsheetml/2009/9/main}"
    ns_xm = "{http://schemas.microsoft.com/office/excel/2006/main}"

    results: list[tuple[str, str]] = []
    with ZipFile(workbook_path) as z:
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))

        rid_to_target: dict[str, str] = {}
        for rel in rels:
            if rel.tag.endswith("Relationship"):
                rid_to_target[rel.attrib.get("Id", "")] = rel.attrib.get("Target", "")

        sheet_xml_path = ""
        for s in wb.findall(f".//{ns_main}sheet"):
            if s.attrib.get("name") != sheet_name:
                continue
            rid = s.attrib.get(f"{ns_rel}id", "")
            target = rid_to_target.get(rid, "")
            if target:
                sheet_xml_path = "xl/" + target
            break

        if not sheet_xml_path:
            return results

        root = ET.fromstring(z.read(sheet_xml_path))
        for dv in root.findall(f".//{ns_x14}dataValidation"):
            dv_type = dv.attrib.get("type", "")
            if dv_type != "list":
                continue

            formula_node = dv.find(f"{ns_x14}formula1/{ns_xm}f")
            sqref_node = dv.find(f"{ns_xm}sqref")
            formula = "" if formula_node is None else (formula_node.text or "")
            sqref = "" if sqref_node is None else (sqref_node.text or "")
            if formula and sqref:
                results.append((sqref.strip(), formula.strip()))

    return results


def restore_x14_dropdowns_from_reference(target_workbook, reference_path: Path) -> int:
    rules = _extract_x14_dropdown_validations(reference_path, "Schutzprüf-Checkliste")
    if not rules:
        return 0

    ws = target_workbook.Worksheets("Schutzprüf-Checkliste")
    try:
        ws.Unprotect()
    except Exception:
        pass

    restored = 0
    for sqref, formula in rules:
        try:
            # COM range union syntax uses commas instead of spaces.
            target_ref = sqref.replace(" ", ",")
            rng = ws.Range(target_ref)
            try:
                rng.Validation.Delete()
            except Exception:
                pass

            if not formula.startswith("="):
                formula = "=" + formula

            # xlValidateList = 3, xlValidAlertStop = 1
            rng.Validation.Add(Type=3, AlertStyle=1, Formula1=formula)
            rng.Validation.IgnoreBlank = True
            rng.Validation.InCellDropdown = True
            rng.Validation.ShowInput = True
            rng.Validation.ShowError = True
            restored += 1
        except Exception:
            continue

    return restored


def ensure_checkliste_e_column_dropdown_fallback(workbook) -> int:
    ws = workbook.Worksheets("Schutzprüf-Checkliste")
    try:
        ws.Unprotect()
    except Exception:
        pass

    # Typical checklist input area where x/!/ ? are entered.
    added_cells = 0
    for row in range(3, 401):
        cell = ws.Cells(row, 5)  # Column E
        try:
            if bool(cell.MergeCells):
                continue
        except Exception:
            pass

        try:
            if int(cell.Validation.Type) != 0:
                continue
        except Exception:
            pass

        try:
            # xlValidateList = 3, xlValidAlertStop = 1
            cell.Validation.Add(Type=3, AlertStyle=1, Formula1='"x,!,?,-"')
            cell.Validation.IgnoreBlank = True
            cell.Validation.InCellDropdown = True
            cell.Validation.ShowInput = True
            cell.Validation.ShowError = True
            cell.Validation.InputTitle = "Eingabe"
            cell.Validation.InputMessage = "Bitte x, !, ? oder - auswählen."
            cell.Validation.ErrorTitle = "Ungültige Eingabe"
            cell.Validation.ErrorMessage = "Nur x, !, ? oder - sind erlaubt."
            added_cells += 1
        except Exception:
            continue

    return added_cells


def ensure_aktuelles_datum_macro(workbook) -> int:
    vbcomponents = workbook.VBProject.VBComponents
    module_name = "modCopilotDateFix"

    try:
        component = vbcomponents(module_name)
    except Exception:
        # 1 = vbext_ct_StdModule
        component = vbcomponents.Add(1)
        component.Name = module_name

    code_module = component.CodeModule
    if code_module.CountOfLines > 0:
        code_module.DeleteLines(1, code_module.CountOfLines)
    code_module.AddFromString(DATE_MACRO_CODE)

    rebound = 0
    macro_ref = f"'{workbook.Name}'!AktuellesDatum"

    for ws in workbook.Worksheets:
        for i in range(1, int(ws.Shapes.Count) + 1):
            shape = ws.Shapes(i)
            text_blob = ""

            try:
                text_blob += " " + str(shape.TextFrame.Characters().Text)
            except Exception:
                pass

            try:
                text_blob += " " + str(shape.TextFrame2.TextRange.Text)
            except Exception:
                pass

            try:
                text_blob += " " + str(shape.OnAction)
            except Exception:
                pass

            if "datum" not in text_blob.lower():
                continue

            try:
                shape.OnAction = macro_ref
                rebound += 1
            except Exception:
                continue

    return rebound


def backup_file(path: Path, suffix: str) -> Path:
    backup_path = path.with_name(f"{path.stem}{suffix}{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def norm_name(value: str) -> str:
    text = value.replace("\r", "\n")
    first = text.split("\n", 1)[0].strip().lower()
    repl = {
        "ae": "a",
        "oe": "o",
        "ue": "u",
        "ss": "s",
        "gmbh": "",
        "co. kg": "",
        "co kg": "",
        "kg": "",
        "ag": "",
        "-": " ",
        ",": " ",
        ".": " ",
    }
    for old, new in repl.items():
        first = first.replace(old, new)
    while "  " in first:
        first = first.replace("  ", " ")
    return first.strip()


def is_placeholder_email(value: str) -> bool:
    lower = value.strip().lower()
    if lower in {"", "?", "-"}:
        return True
    if "genauerausw" in lower:
        return True
    if "ueber " in lower:
        return True
    return False


def load_source_emails(source_path: Path) -> list[tuple[str, str, str]]:
    wb = load_workbook(source_path, data_only=True)
    ws = None
    for name in SOURCE_CANDIDATE_SHEETS:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        raise RuntimeError(
            "Kein Blatt 'Kundenadressen' oder 'Kunden' in der Quelle gefunden."
        )

    result: list[tuple[str, str, str]] = []
    for row in range(1, ws.max_row + 1):
        customer = ws.cell(row, 1).value
        email = ws.cell(row, 2).value
        if customer is None:
            continue
        customer_text = str(customer).strip()
        if not customer_text or customer_text.lower() == "kunde":
            continue
        email_text = "" if email is None else str(email).strip()
        if is_placeholder_email(email_text):
            continue
        result.append((norm_name(customer_text), customer_text, email_text))
    return result


def pick_email_for_customer(customer_value: str, source_rows: list[tuple[str, str, str]]) -> str:
    target_key = norm_name(customer_value)
    if not target_key:
        return ""

    exact = [email for key, _raw, email in source_rows if key == target_key]
    if exact:
        return "; ".join(dict.fromkeys(exact))

    fuzzy = [
        email
        for key, _raw, email in source_rows
        if key and (target_key in key or key in target_key)
    ]
    if fuzzy:
        return "; ".join(dict.fromkeys(fuzzy))

    return ""


def update_customer_sheet(path: Path, source_rows: list[tuple[str, str, str]]) -> tuple[int, int]:
    wb = load_workbook(path, keep_vba=True)
    if CUSTOMER_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Blatt '{CUSTOMER_SHEET}' nicht gefunden in {path}.")
    ws = wb[CUSTOMER_SHEET]

    updated = 0
    missing = 0
    for row in range(1, ws.max_row + 1):
        customer = ws.cell(row, 1).value
        if customer is None:
            continue
        customer_text = str(customer).strip()
        if not customer_text:
            continue

        email = pick_email_for_customer(customer_text, source_rows)
        if email:
            if ws.cell(row, 2).value != email:
                ws.cell(row, 2).value = email
                updated += 1
        else:
            ws.cell(row, 2).value = ""
            missing += 1

    wb.save(path)
    wb.close()
    return updated, missing


def update_customer_sheet_excel_com(workbook, source_rows: list[tuple[str, str, str]]) -> tuple[int, int]:
    try:
        ws = workbook.Worksheets(CUSTOMER_SHEET)
    except Exception as error:
        raise RuntimeError(f"Blatt '{CUSTOMER_SHEET}' nicht gefunden in {workbook.Name}.") from error

    last_row = int(ws.Cells(ws.Rows.Count, 1).End(-4162).Row)  # xlUp
    updated = 0
    missing = 0

    for row in range(1, last_row + 1):
        customer_value = ws.Cells(row, 1).Value
        if customer_value is None:
            continue
        customer_text = str(customer_value).strip()
        if not customer_text:
            continue

        email = pick_email_for_customer(customer_text, source_rows)
        current_email = ws.Cells(row, 2).Value
        current_text = "" if current_email is None else str(current_email)

        if email:
            if current_text != email:
                ws.Cells(row, 2).Value = email
                updated += 1
        else:
            if current_text != "":
                ws.Cells(row, 2).Value = ""
            missing += 1

    return updated, missing


def _extract_com_hresult(error: Exception) -> int | None:
    if not getattr(error, "args", None):
        return None
    first_arg = error.args[0]
    if isinstance(first_arg, int):
        return first_arg
    return None


def _is_excel_call_rejected(error: Exception) -> bool:
    # RPC_E_CALL_REJECTED
    return _extract_com_hresult(error) == -2147418111


def _retry_excel_call(callable_fn, attempts: int = 20, wait_seconds: float = 0.35):
    for attempt in range(1, attempts + 1):
        try:
            return callable_fn()
        except Exception as error:
            if (not _is_excel_call_rejected(error)) or attempt == attempts:
                raise
            time.sleep(wait_seconds)


def apply_abschlussbemerkungen_nicht_ok_logic(workbook) -> None:
    ws_angaben = workbook.Worksheets("Allgemeine Angaben")
    ws_protokoll = workbook.Worksheets("Prüfprotokoll")

    # Map "!" from checklist-linked fields to "NICHT OK" so existing warning semantics apply.
    formula_updates = {
        "C105": "=IF('Schutzprüf-Checkliste'!I70=TRUE,IF(OR('Schutzprüf-Checkliste'!E70=\"!\",'Schutzprüf-Checkliste'!J70=\"!\"),\"NICHT OK\",\"x\"),\"\")",
        "C106": "=IF('Schutzprüf-Checkliste'!K70=TRUE,IF(OR('Schutzprüf-Checkliste'!E70=\"!\",'Schutzprüf-Checkliste'!L70=\"!\"),\"NICHT OK\",\"x\"),\"\")",
        "C107": "=IF('Schutzprüf-Checkliste'!I71=TRUE,IF(OR('Schutzprüf-Checkliste'!E71=\"!\",'Schutzprüf-Checkliste'!J71=\"!\"),\"NICHT OK\",\"x\"),\"\")",
        "C108": "=IF('Schutzprüf-Checkliste'!K71=TRUE,IF(OR('Schutzprüf-Checkliste'!E71=\"!\",'Schutzprüf-Checkliste'!L71=\"!\"),\"NICHT OK\",\"x\"),\"\")",
        "C114": "=IF('Schutzprüf-Checkliste'!I72=TRUE,IF(OR('Schutzprüf-Checkliste'!E72=\"!\",'Schutzprüf-Checkliste'!J72=\"!\"),\"NICHT OK\",\"x\"),\"\")",
        "C115": "=IF('Schutzprüf-Checkliste'!K72=TRUE,IF(OR('Schutzprüf-Checkliste'!E72=\"!\",'Schutzprüf-Checkliste'!L72=\"!\"),\"NICHT OK\",\"x\"),\"\")",
        "C109": "=IF('Schutzprüf-Checkliste'!E74=\"x\",\"x\",IF('Schutzprüf-Checkliste'!E74=\"!\",\"NICHT OK\",\"\"))",
    }
    for cell, formula in formula_updates.items():
        ws_angaben.Range(cell).Formula = formula

    # Ensure description rows in Abschlussbemerkungen are also shown for "!".
    protokoll_formula_updates = {
        "A169": "=IF(AND('Schutzprüf-Checkliste'!I70=TRUE,OR(J169=\"x\",J169=\"NICHT OK\")),\"Abschaltung MS-LS\"&'Schutzprüf-Checkliste'!C67 &\" nach Ausfall der Hilfsspannung - AuxDC\",\"\")",
        "A170": "=IF(AND('Schutzprüf-Checkliste'!K70=TRUE,OR(J170=\"x\",J170=\"NICHT OK\")),\"Abschaltung NS-LS\"&'Schutzprüf-Checkliste'!C68 &\" nach Ausfall der Hilfsspannung- AuxDC\",\"\")",
        "A171": "=IF(AND('Schutzprüf-Checkliste'!I71=TRUE,OR(J171=\"x\",J171=\"NICHT OK\")),\"Abschaltung MS-LS\"&'Schutzprüf-Checkliste'!C67 &\" nach Ausfall Schutzrelais (Live Contact)\",\"\")",
        "A172": "=IF(AND('Schutzprüf-Checkliste'!K71=TRUE,OR(J172=\"x\",J172=\"NICHT OK\")),\"Abschaltung NS-LS\"&'Schutzprüf-Checkliste'!C68 &\" nach Ausfall Schutzrelais (Live Contact)\",\"\")",
        "A173": "=IF(AND('Schutzprüf-Checkliste'!I72=TRUE,OR(J173=\"x\",J173=\"NICHT OK\")),'Allgemeine Angaben'!A114,\"\")",
        "A174": "=IF(AND('Schutzprüf-Checkliste'!K72=TRUE,OR(J174=\"x\",J174=\"NICHT OK\")),'Allgemeine Angaben'!A115,\"\")",
        "A182": "=IF(OR('Schutzprüf-Checkliste'!E74=\"x\",'Schutzprüf-Checkliste'!E74=\"!\"),\"DC USV für übergeordneter Schutz/ggfs. UMZ-Schutz in Ordnung\",\"\")",
    }
    for cell, formula in protokoll_formula_updates.items():
        ws_protokoll.Range(cell).Formula = formula

    # Existing red rule covers J168:J175. Add same semantic rule for J182 (linked to row 74).
    try:
        target_cell = ws_protokoll.Range("J182")
        has_nicht_ok_rule = False
        for i in range(1, int(target_cell.FormatConditions.Count) + 1):
            rule = target_cell.FormatConditions(i)
            try:
                if int(rule.Type) == 1 and int(rule.Operator) == 3:
                    formula = str(rule.Formula1).replace("=", "").replace('"', "").strip().upper()
                    if formula == "NICHT OK":
                        has_nicht_ok_rule = True
                        break
            except Exception:
                continue

        if not has_nicht_ok_rule:
            added_rule = target_cell.FormatConditions.Add(Type=1, Operator=3, Formula1='="NICHT OK"')
            try:
                ref_rule = ws_protokoll.Range("J168").FormatConditions(1)
                added_rule.Font.Color = ref_rule.Font.Color
                added_rule.Interior.Color = ref_rule.Interior.Color
                added_rule.StopIfTrue = ref_rule.StopIfTrue
            except Exception:
                # Style copy is best effort; semantic condition is the functional requirement.
                pass
    except Exception as error:
        print(f"Hinweis: CF-Regel fuer J182 konnte nicht gesetzt/geprueft werden: {error}")

    # Show station lock message in row 3 when any lower status is NICHT OK.
    lock_message = "Station gesperrt - Nicht alle Schutzfunktionen in Ordnung"
    lock_formula = (
        "=IF(OR(COUNTIF($J$167:$J$175,\"NICHT OK\")>0,$J$182=\"NICHT OK\",$J$169=\"NICHT OK\",$J$170=\"NICHT OK\",$J$171=\"NICHT OK\",$J$172=\"NICHT OK\",$J$173=\"NICHT OK\",$J$174=\"NICHT OK\"),"
        f"\"{lock_message}\""
        ",\"Prüfer\")"
    )
    ws_protokoll.Range("A3").Formula = lock_formula

    lock_cf_formula = f'="{lock_message}"'
    try:
        a3_cell = ws_protokoll.Range("A3")
        a3_cell.FormatConditions.Delete()
        lock_rule = a3_cell.FormatConditions.Add(Type=1, Operator=3, Formula1=lock_cf_formula)
        lock_rule.Interior.Color = 255
        lock_rule.Font.Color = 16777215
    except Exception as error:
        print(f"Hinweis: CF-Regel fuer A3 konnte nicht gesetzt/geprueft werden: {error}")


def get_abschlussbemerkungen_debug_snapshot(workbook) -> dict[str, str]:
    ws_angaben = workbook.Worksheets("Allgemeine Angaben")
    ws_protokoll = workbook.Worksheets("Prüfprotokoll")
    keys = {
        "C105": ws_angaben.Range("C105").Formula,
        "C106": ws_angaben.Range("C106").Formula,
        "C107": ws_angaben.Range("C107").Formula,
        "C108": ws_angaben.Range("C108").Formula,
        "C109": ws_angaben.Range("C109").Formula,
        "C114": ws_angaben.Range("C114").Formula,
        "C115": ws_angaben.Range("C115").Formula,
        "A169": ws_protokoll.Range("A169").Formula,
        "A170": ws_protokoll.Range("A170").Formula,
        "A171": ws_protokoll.Range("A171").Formula,
        "A172": ws_protokoll.Range("A172").Formula,
        "A173": ws_protokoll.Range("A173").Formula,
        "A174": ws_protokoll.Range("A174").Formula,
        "A182": ws_protokoll.Range("A182").Formula,
        "A3": ws_protokoll.Range("A3").Formula,
    }
    return {k: str(v) for k, v in keys.items()}


def patch_pdf_form_vba(path: Path, visible: bool) -> None:
    try:
        import pythoncom
        import win32com.client
    except ImportError as error:
        raise RuntimeError(
            "Excel-COM ist nicht verfuegbar. Bitte unter Windows mit pywin32 ausfuehren."
        ) from error

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = bool(visible)
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        workbook = excel.Workbooks.Open(str(path.resolve()))
        component = workbook.VBProject.VBComponents("frmPDFDruck")
        module = component.CodeModule

        if module.CountOfLines > 0:
            module.DeleteLines(1, module.CountOfLines)
        module.AddFromString(FORM_CODE)

        workbook.Save()
        workbook.Close(SaveChanges=True)
        workbook = None
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def apply_changes_with_excel_com(
    path: Path,
    source_rows: list[tuple[str, str, str]],
    visible: bool,
    dropdown_reference_path: Path | None = None,
) -> tuple[int, int]:
    try:
        import pythoncom
        import win32com.client
        import win32api
    except ImportError as error:
        raise RuntimeError(
            "Excel-COM ist nicht verfuegbar. Bitte unter Windows mit pywin32 ausfuehren."
        ) from error

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    temp_dir: str | None = None
    temp_workbook_path: Path | None = None
    original_path = path.resolve()
    debug_snapshot: dict[str, str] = {}

    def _open_workbook(excel_app, workbook_path_str: str):
        last_error: Exception | None = None
        # Try normal open first.
        open_variants = [
            {},
            {
                "UpdateLinks": 0,
                "ReadOnly": False,
                "IgnoreReadOnlyRecommended": True,
                "AddToMru": False,
                "Local": True,
            },
            {
                "UpdateLinks": 0,
                "ReadOnly": False,
                "IgnoreReadOnlyRecommended": True,
                "AddToMru": False,
                "Local": True,
                # 1 = xlRepairFile
                "CorruptLoad": 1,
            },
        ]

        for kwargs in open_variants:
            try:
                if kwargs:
                    return _retry_excel_call(lambda kwargs=kwargs: excel_app.Workbooks.Open(workbook_path_str, **kwargs))
                return _retry_excel_call(lambda: excel_app.Workbooks.Open(workbook_path_str))
            except Exception as error:
                last_error = error
                continue

        if last_error is not None:
            raise last_error
        raise RuntimeError("Unbekannter Fehler beim Oeffnen der Excel-Datei.")

    try:
        excel = _retry_excel_call(lambda: win32com.client.DispatchEx("Excel.Application"))
        excel.Visible = bool(visible)
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        workbook_path = str(original_path)
        try:
            short_path = win32api.GetShortPathName(workbook_path)
            if short_path:
                workbook_path = short_path
        except Exception:
            pass

        try:
            workbook = _open_workbook(excel, workbook_path)
        except Exception:
            # Fallback: copy to temp path with ASCII-only filename and retry there.
            temp_dir = tempfile.mkdtemp(prefix="ges_excel_open_")
            temp_workbook_path = Path(temp_dir) / "workbook.xlsm"
            shutil.copy2(original_path, temp_workbook_path)
            workbook = _open_workbook(excel, str(temp_workbook_path))

        print(f"{path}: Schritt Kundenblatt aktualisieren ...")
        updated, missing = _retry_excel_call(lambda: update_customer_sheet_excel_com(workbook, source_rows))

        print(f"{path}: Schritt Abschlussbemerkungen/NICHT-OK-Logik ...")
        try:
            _retry_excel_call(lambda: apply_abschlussbemerkungen_nicht_ok_logic(workbook))
        except Exception as error:
            raise RuntimeError(
                "Fehler in Schritt 'Abschlussbemerkungen/NICHT-OK-Logik'. "
                f"Original: {error}"
            ) from error

        print(f"{path}: Schritt Debug-Snapshot ...")
        try:
            debug_snapshot = _retry_excel_call(lambda: get_abschlussbemerkungen_debug_snapshot(workbook))
        except Exception as error:
            # Snapshot is diagnostic only and must not abort processing.
            debug_snapshot = {}
            print(f"{path}: Hinweis: Debug-Snapshot konnte nicht gelesen werden: {error}")

        reference_logo_path = resolve_logo_reference_path(path)
        dropdown_reference = dropdown_reference_path
        if dropdown_reference is None or not dropdown_reference.exists():
            dropdown_reference = reference_logo_path
        restored_logos = 0
        restored_dropdown_areas = 0
        restored_x14_dropdowns = 0
        added_dropdown_fallback_cells = 0
        rebound_datum_buttons = 0
        if dropdown_reference is not None:
            print(f"{path}: Schritt Drop-downs aus V19 wiederherstellen ...")
            restored_dropdown_areas = _retry_excel_call(
                lambda: restore_checkliste_dropdowns_from_reference(excel, workbook, dropdown_reference)
            )
            restored_x14_dropdowns = _retry_excel_call(
                lambda: restore_x14_dropdowns_from_reference(workbook, dropdown_reference)
            )
            added_dropdown_fallback_cells = _retry_excel_call(
                lambda: ensure_checkliste_e_column_dropdown_fallback(workbook)
            )

            print(f"{path}: Drop-down-Referenz: {dropdown_reference}")

        if reference_logo_path is not None:

            restored_logos = _retry_excel_call(
                lambda: restore_logos_from_reference(excel, workbook, reference_logo_path)
            )
            print(f"{path}: Logos aus V19 wiederhergestellt (Anzahl: {restored_logos})")
            print(f"{path}: Drop-down-Bereiche in Schutzprüf-Checkliste wiederhergestellt: {restored_dropdown_areas}")
            print(f"{path}: x14-Drop-down-Regeln wiederhergestellt: {restored_x14_dropdowns}")
            print(f"{path}: Drop-down-Fallback E3:E400 ergänzt (Zellen): {added_dropdown_fallback_cells}")
        else:
            if dropdown_reference is not None:
                print(f"{path}: Keine V19-Referenzdatei fuer Logo-Wiederherstellung gefunden")
                print(f"{path}: Drop-down-Bereiche in Schutzprüf-Checkliste wiederhergestellt: {restored_dropdown_areas}")
                print(f"{path}: x14-Drop-down-Regeln wiederhergestellt: {restored_x14_dropdowns}")
                print(f"{path}: Drop-down-Fallback E3:E400 ergänzt (Zellen): {added_dropdown_fallback_cells}")
            else:
                print(f"{path}: Keine Referenzdatei fuer Drop-down- oder Logo-Wiederherstellung gefunden")

        print(f"{path}: Schritt Makro 'Aktuelles Datum' robust setzen ...")
        rebound_datum_buttons = _retry_excel_call(lambda: ensure_aktuelles_datum_macro(workbook))
        print(f"{path}: Buttons auf Makro AktuellesDatum gebunden: {rebound_datum_buttons}")

        component = _retry_excel_call(lambda: workbook.VBProject.VBComponents("frmPDFDruck"))
        module = component.CodeModule
        if module.CountOfLines > 0:
            _retry_excel_call(lambda: module.DeleteLines(1, module.CountOfLines))
        _retry_excel_call(lambda: module.AddFromString(FORM_CODE))

        _retry_excel_call(lambda: workbook.Save())
        _retry_excel_call(lambda: workbook.Close(SaveChanges=True))
        workbook = None

        if temp_workbook_path is not None:
            shutil.copy2(temp_workbook_path, original_path)

        print(f"{path}: Abschlussbemerkungen-Formeln aktualisiert")
        if debug_snapshot:
            for key, value in debug_snapshot.items():
                print(f"{path}: {key} -> {value}")

        return updated, missing
    finally:
        if workbook is not None:
            _retry_excel_call(lambda: workbook.Close(SaveChanges=False))
        if excel is not None:
            _retry_excel_call(lambda: excel.Quit())
        pythoncom.CoUninitialize()
        if temp_dir is not None:
            shutil.rmtree(temp_dir, ignore_errors=True)


def main() -> int:
    args = parse_args()
    source_path = resolve_source_path(args.source)

    requested_targets = [Path(p) for p in args.targets]
    targets = resolve_target_paths(requested_targets)

    source_rows = load_source_emails(source_path)
    if not source_rows:
        raise RuntimeError("Keine nutzbaren Kunde->E-Mail-Daten in der Quelle gefunden.")

    print(f"Quelle geladen: {source_path}")
    print(f"Nutzbare E-Mail-Zuordnungen: {len(source_rows)}")
    print("Ziel-Dateien:")
    for target in targets:
        print(f"- {target}")

    failures: list[str] = []

    for target in targets:
        try:
            backup_path = backup_file(target, args.backup_suffix)
            print(f"Sicherung erstellt: {backup_path}")

            updated, missing = apply_changes_with_excel_com(
                target,
                source_rows,
                visible=args.visible,
                dropdown_reference_path=backup_path,
            )
            print(f"{target}: Kundenblatt aktualisiert (gesetzt: {updated}, ohne Treffer: {missing})")
            print(f"{target}: VBA frmPDFDruck aktualisiert")
        except Exception as error:
            failures.append(f"{target}: {error}")
            print(f"Warnung: Verarbeitung fehlgeschlagen fuer {target}: {error}")

    if failures:
        print("\nFolgende Dateien konnten nicht verarbeitet werden:", file=sys.stderr)
        for item in failures:
            print(f"- {item}", file=sys.stderr)
        return 1

    print("Fertig: PDF->E-Mail-Option und Kunden-E-Mail-Spalte sind eingepflegt.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"Fehler: {exc}", file=sys.stderr)
        raise SystemExit(1)
